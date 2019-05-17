#!python
# Uses Python 3
# needs to install openpyxl

from openpyxl import load_workbook
import re

wb = load_workbook("./parse.xlsx")

print(wb.sheetnames)

for sheet in wb:
  if sheet.title == "GlossaryNotes":
    continue
  # Take sheet name, trim the 'n - ' (Where n is a number) and set it as character name
  print("Splitting: " + sheet.title)
  charName = re.split(r"[a-zA-Z]+", sheet.title)[0]
  # From each line I have: 
  # Column | Desc
  # 0 move name
  # 1 Startup, 2 Total Frames, Landing lag
  # 5 Base dmg, 6 shieldlag, 7 shieldstun, 8 which hitbox, 9 Advantage
  for line in sheet:
    if line[0].value != None:
      print(line[0].value)